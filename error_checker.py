
import glob
import re
import csv
import pandas as pd
from openpyxl import load_workbook
import xlrd
import xlwings as xw
import xlsxwriter
import win32com.client as win32
import pywin
import os
import docx
import parser_functions as pf
import case_finder
from docx import Document

## NEED TO FIGURE OUT HOW TO OPEN PASS-PROTECTED FILE

#xlApp = win32.Dispatch("Excel.Application")

""" Make this interact with, or draw inspiration from,
the make_new_rows() fxn in parser_functions.py """
parser_path = 'R:/groups/seeley/Mack/NP report parser/'

## NOTE: make it so that if there are multiple DDS files, offer choice; otherwise, default to the one in the folder
DDS = 'R:/groups/seeley_pathology/LAVA/DIAGNOSTIC DATA UPLOAD_20190429.xlsx'
HS_data = r'R:/groups/seeley/Mack/NP report parser/HS Discrepancy Data.xlsx'
HS_data = r'R:\groups\seeley\Mack\NP report parser\HS Discrepancy Data.xlsx'
error_checking_sheet = r'R:/groups/seeley/Mack/NP report parser/error checking sheet.xlsx'
microinfarcts_file = r'R:/groups/seeley/Mack/NP report parser/microinfarct_phrases_2.xlsx'


#wb = xlApp.Workbooks.Open(DDS, False, True, None, 'spindle4')
#wb = xw.Book(DDS)
#sheet = wb.sheets['2017 data fields conditional']
#df = sheet.options(pd.DataFrame, index=False, header=True).value

# Test of grabbing headers -- better to grab them from DDS, though
# Need fxn to look for discrepancies between headers in error-checker
# and headers in DDS
DDS_file = pd.ExcelFile(DDS)
data = pd.read_excel(DDS_file, '2017 data fields conditional')
df = pd.DataFrame(data)
fields = list(df)

# List of cases to check (pNums)
pNums_to_check = pd.DataFrame(pd.read_excel(error_checking_sheet, sheet_name='pNums_to_check'))
pNums_to_check = list(pNums_to_check['P#'])
#pNums_to_check = ['P2802']
print(pNums_to_check)


# Store Primary, Contributing and Incidental Dx column field names from DDS
dx_columns = [key for key in fields if any(x in key for x in ['PrimDx', 'ContribDx', 'IncidDx'])]

# Store white matter rarefaction (WMR) field names
WMR_columns = []
for item in fields:
    if 'WhiteMatterRarefaction' in item:
        WMR_columns.append(item)
#print(WMR_columns)

# Store microbleed field names
microbleed_columns = []
for item in fields:
    if 'Microbleed' in item:
        microbleed_columns.append(item)

# Store microinfarct excel book


# Store microinfarct columns
microinfarct_columns = []
for item in fields:
    if 'Microinfarct' in item:
        microinfarct_columns.append(item)

# Store lacunar infarct columns
lacune_columns = []
for item in fields:
    if 'Lacunar infarct' in item:
        lacune_columns.append(item)

# Store territorial infarct columns
territorial_columns = []
for item in fields:
    if 'Territorial infarct' in item:
        territorial_columns.append(item)

vbi_columns = microbleed_columns + microinfarct_columns + lacune_columns + territorial_columns
#print(vbi_columns)

after_stages = ['Additional observations', 'Pending issues', 'Approved by', 'Finalized date', 'Discrepancies with report']

fields_to_ignore = ['PIDN'] + dx_columns + ['TDPProteinopathynonFTLD', 'CTE', 'HippocampalSclerosis',
                                            'HippocampalSclerosis_Laterality',
                                            'SubduralHematoma_laterality'] + WMR_columns + microbleed_columns + lacune_columns + territorial_columns + after_stages
print(fields_to_ignore)

working_fields = []

for item in fields:
    if item in fields_to_ignore:
        continue
    else:
        working_fields.append(item)

print(f'WORKING FIELDS: {working_fields}')
print(len(working_fields))

# Returns dictionary of all specific regions identified in microinfarcts section with their 'regional category.'
# Key: region, Value: category (ALL CAPS)
# Put in parser_functions
def microinfarct_region2cat():
    region_worksheet = pd.read_excel(parser_path + 'microinfarct_phrases_2.xlsx', sheet_name='regions')
    reg_df_0 = pd.DataFrame(region_worksheet)
    reg_df = reg_df_0[['Region', 'RegDesignation_1']]
    zipped_list = list(zip(reg_df.Region, reg_df.RegDesignation_1))
    region2cat = dict((x, y) for x, y in zipped_list)
    print(region2cat)
    return region2cat

def microinfarct_phrase2score():
    phrase_worksheet = pd.read_excel(microinfarcts_file, sheet_name='distinct sentences')
    working_columns = ['Sentence']
    for key in microinfarct_columns:
        working_columns.append(key)
    phrase_df_0 = pd.DataFrame(phrase_worksheet)
    phrase_df = phrase_df_0[working_columns]
    phrase_scores_list = phrase_df.values.tolist()
    fields = list(phrase_df)
    sentence_dict = {}
    for row in phrase_scores_list:
        sentence = row[0]
        sentence_dict[sentence] = row[1:]
    print(sentence_dict)
    return sentence_dict

    # colsIwant = [c for c in phrase_worksheet.columns if 'Microinfarct' in c]
    #phrase_df_0 = phrase_worksheet[colsIwant]

def add_microinfarct_score(sentences): #First, test whether we can successfully add the scores...
    sentence_dict = microinfarct_phrase2score()

    sentence_scores = []
    key_errors = []

    for sentence in sentences:
        try:
            new_row = sentence_dict[sentence]
            sentence_scores.append(new_row)
        except KeyError:
            print(f'Key error: {sentence}')
            key_errors.append(sentence)
            continue
    df = pd.DataFrame(sentence_scores, columns=microinfarct_columns)
    df.loc['Total'] = df.sum()
    print(df)
    if key_errors:
        print(f'Key errors:')
        for item in key_errors:
            print(item)
    total_scores = list(df.loc['Total'])
    total_scores = [score if score <= 3 else 3 for score in total_scores]
    df.loc['Total'] = total_scores
    return total_scores

def check_microinfarcts(pNum):
    dict_entry = pf.get_files([pNum])
    filename = dict_entry[pNum]
    #print(f'Filename: {filename}')
    lines = pf.open_file(filename)

    microinfarct_sxn = case_finder.get_microinfarct_section(lines)
    lines = []
    for line in microinfarct_sxn:
        if 'FOUND' in line:
            lines.append(line)

    distinct_sentences_list = []

    for line in lines:
        distinct_sentences = line.split('.')
        for sentence in distinct_sentences:
            sentence = sentence.replace('FINDINGS:', '')
            sentence = sentence.replace(r'\t', '')
            sentence = sentence.strip()
            if sentence == '':
                continue
            distinct_sentences_list.append(sentence)
    #print(f'{pNum} sentences: {distinct_sentences_list}')

    values = add_microinfarct_score(distinct_sentences_list)
    microinfarct_values = dict(zip(microinfarct_columns, values))
    #for key, value in microinfarct_values.items():
    #    print(f'{key}: {value}')
    return list(microinfarct_values.values())


def parser_rows():
    file_dict = pf.get_files(pNums_to_check)

    # Set working fields to column names spanning entire DDS


    # Create empty dataframe with parser fields as columns
    data = []


    for pNum in file_dict.keys():
        print(f' --- {pNum} ---')
        filename = file_dict[pNum]
        all_contents = pf.open_file(filename)
        dx_sxn = pf.get_dx_sxn(filename)
        grossObs = pf.get_grossObs(filename)
        site = 'UCSF NDBB'
        author = pf.get_author(pNum)
        ADNC_dict = pf.get_ADNC(dx_sxn)
        Thal_phase = ADNC_dict['Thal Phase']
        AD_Braak = ADNC_dict['Braak Stage']
        AD_CERAD_NP = ADNC_dict['CERAD NP Score']
        AD_CERAD_DP = ADNC_dict['CERAD DP Score']
        NIAReag = ADNC_dict['NIA-Reagan']
        CAA = pf.get_CAA(dx_sxn)
        ADNC_level = ADNC_dict['ADNC level']
        LBD = pf.get_lbd_stage(dx_sxn)
        PD_Braak = pf.get_PDBraak(dx_sxn)
        ATAC = pf.get_ATAC(dx_sxn)
        #CTE = pf.get_CTE(dx_sxn)
        #HS = 'NA'
        #HS_laterality = 'NA'
        Arterio = pf.get_arterio(dx_sxn)
        Athero = pf.get_athero(grossObs)
        #TDP_proteinopathy = 'NA'
        AGD = pf.get_AGD(dx_sxn)
        HD = pf.get_huntington(dx_sxn)
        microinfarcts = check_microinfarcts(pNum)

        parser_values = [pNum, site, author, Thal_phase, AD_Braak, AD_CERAD_NP, AD_CERAD_DP, NIAReag, CAA, ADNC_level, LBD, PD_Braak,
                          ATAC, Arterio, Athero, AGD, HD] + microinfarcts
        data.append(parser_values)

        working_data = dict(zip(working_fields, parser_values))
        print(working_data)

        """
        primDx_list = pf.get_PrimDx(dx_sxn)
        print(f'\nNumber of primary diagnoses: {len(primDx_list)}')
        for counter, dx in enumerate(primDx_list, 1):
            print(f'{counter}: {dx}')

        contributingDx_list = pf.get_ContributingDx(dx_sxn)
        print(f'\nNumber of contributing diagnoses: {len(contributingDx_list)}')
        for counter, dx in enumerate(contributingDx_list, 1):
            print(f'{counter}: {dx}')

        incidentalDx_list = pf.get_IncidentalDx(dx_sxn)
        print(f'\nNumber of incidental diagnoses: {len(incidentalDx_list)}')
        for counter, dx in enumerate(incidentalDx_list, 1):
            print(f'{counter}: {dx}')
        """
    working_df = pd.DataFrame(data, columns=working_fields)
    #print(working_df)
    book = load_workbook(error_checking_sheet)
    with pd.ExcelWriter(error_checking_sheet, engine='openpyxl') as writer:
        writer.book = book
        working_df.to_excel(writer, 'parser values')
    return working_df

"""
        parser_fields = book.get_sheet_by_name('parser fields')
        for key in working_data.keys():
            parser_value = working_data[key]
            parser_fields[key] = parser_value
"""
#def set_rows(df):

def compare_columns():
    parser_df = parser_rows()
    dds_df = pd.read_excel(DDS, sheet_name='2017 data fields conditional')
    #dds_df.set_index('P#', inplace=True)


    dds_abridged0 = dds_df[working_fields]
    dds_abridged2 = dds_abridged0.loc[dds_abridged0['P#'].isin(pNums_to_check)]
    #print(dds_abridged2)




    dds_rows = []
    for index, row in dds_abridged2.iterrows():
        dds_row = list(row)
        dds_rows.append(dds_row)
        dds_rows = sorted(dds_rows)

    for row in dds_rows:
        print(row)

    p_rows = []
    for index, row in parser_df.iterrows():
        parser_row = list(row)
        p_rows.append(parser_row)
        p_rows = sorted(p_rows)

    for row in p_rows:
        print(row)

    error_rows = []

    i = 0
    for row in p_rows:
        pNum1 = row[0]
        pNum2 = dds_rows[i][0]

        if pNum1 != pNum2:
            print(pNum1, pNum2)
            continue
        else:
            dds_data = dds_rows[i]
            if dds_data == row:
                error_rows.append(row)
            else:
                error_row = []
                j = 0
                for item in row:
                    if dds_data[j] == item:
                        error_row.append(item)
                    else:
                        print(f'---\nParser {pNum1}\n{working_fields[j]}: {item}\nDDS {pNum2}\n{working_fields[j]}: {dds_data[j]}\n---')
                        error_row.append('Mismatch!')
                    j +=1
                error_rows.append(error_row)
        i += 1
    errors_df = []
    for row in error_rows:
        errors_df.append(row)
    errors_df = pd.DataFrame(errors_df, columns=working_fields)
    print(errors_df)
    return(errors_df)

error_checker = compare_columns()

book = load_workbook(error_checking_sheet)
with pd.ExcelWriter(error_checking_sheet, engine='openpyxl') as writer:
    writer.book = book
    error_checker.to_excel(writer, 'error check')

#check_microinfarcts()









    # Abridge DDS
    #for pNum in pNums_to_check:
    #    for row in dds_df.iterrows():
    #        if row[1] == pNum:
    #            print(row)
    #dds_abridged = dds_df.loc[pNums_to_check, working_fields]
    #print(dds_abridged)
    #for key in working_fields:


#parser_rows()
compare_columns()

